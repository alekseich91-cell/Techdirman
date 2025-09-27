"""
Назначение: Импорт .xlsx с диалогом сопоставления столбцов.
"""
from PySide6 import QtWidgets
from openpyxl import load_workbook
from pathlib import Path

class MappingDialog(QtWidgets.QDialog):
    def __init__(self, headers: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Сопоставление столбцов")
        layout = QtWidgets.QFormLayout(self)
        self.box_name = QtWidgets.QComboBox()
        self.box_qty = QtWidgets.QComboBox()
        self.box_coeff = QtWidgets.QComboBox()
        self.box_amount = QtWidgets.QComboBox()
        for b in (self.box_name, self.box_qty, self.box_coeff, self.box_amount):
            b.addItems(["<не использовать>"] + headers)
        self.edit_group = QtWidgets.QLineEdit("Аренда оборудования")
        layout.addRow("Наименование:", self.box_name)
        layout.addRow("Количество:", self.box_qty)
        layout.addRow("Коэффициент:", self.box_coeff)
        layout.addRow("Стоимость (сумма):", self.box_amount)
        layout.addRow("Группа:", self.edit_group)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept); btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def get_mapping(self) -> dict:
        return {"name": self.box_name.currentText(), "qty": self.box_qty.currentText(), "coeff": self.box_coeff.currentText(), "amount": self.box_amount.currentText(), "group": self.edit_group.text().strip() or "Аренда оборудования"}

def import_xlsx(path: Path, parent=None) -> list[dict]:
    wb = load_workbook(filename=path, data_only=True); ws = wb.active
    headers = []; header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_empty) >= max(2, len(row)//3):
            headers = [str(c).strip() if c is not None else "" for c in row]; header_row_idx = i; break
    if not headers: raise RuntimeError("Не удалось найти строку заголовков в Excel.")
    dlg = MappingDialog(headers, parent=parent)
    if dlg.exec() != QtWidgets.QDialog.Accepted: return []
    m = dlg.get_mapping()
    def idx_of(name: str):
        if name == "<не использовать>": return None
        try: return headers.index(name)
        except ValueError: return None
    idx_name = idx_of(m["name"]); idx_qty = idx_of(m["qty"]); idx_coeff = idx_of(m["coeff"]); idx_amount = idx_of(m["amount"])
    group_name = m["group"]
    items: list[dict] = []
    for r in ws.iter_rows(min_row=(header_row_idx or 1)+1, values_only=True):
        vals = ["" if c is None else c for c in r]
        name = (str(vals[idx_name]).strip() if idx_name is not None and idx_name < len(vals) else "")
        if not name: continue
        qty_raw = vals[idx_qty] if idx_qty is not None and idx_qty < len(vals) else 1
        coeff_raw = vals[idx_coeff] if idx_coeff is not None and idx_coeff < len(vals) else 1
        amount_raw = vals[idx_amount] if idx_amount is not None and idx_amount < len(vals) else 0
        def to_float(x):
            if isinstance(x, (int, float)): return float(x)
            s = str(x).replace(" ", "").replace(",", ".")
            try: return float(s)
            except: return 0.0
        qty = max(0.0, to_float(qty_raw)); coeff = max(0.0, to_float(coeff_raw)) or 1.0; amount = max(0.0, to_float(amount_raw))
        unit_price = amount/(qty*coeff) if qty>0 and coeff>0 else 0.0
        items.append({"type":"equipment","group_name":group_name,"name":name,"qty":qty,"coeff":coeff,"amount":amount,"unit_price":unit_price})
    return items
