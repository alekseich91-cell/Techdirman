"""
Назначение:
    Универсальный импорт смет из файлов Excel (XLSX) и CSV с интерактивным сопоставлением столбцов.

Принцип работы:
    1) Определяет тип файла по расширению (.xlsx или .csv).
    2) Для Excel-файлов запрашивает у пользователя лист для чтения (если в книге несколько листов).
    3) Читает заголовки выбранного листа и строит список возможных полей.
    4) Показывает диалог маппинга полей: Наименование, Количество, Коэффициент, Сумма, Подрядчик, Отдел, Класс (РУССКИЙ).
    5) Приводит числа (замена запятой на точку), считает цену за штуку = Сумма/(Кол-во*Коэфф) при валидных множителях.
    6) Возвращает список словарей элементов:
       {
         "name", "qty", "coeff", "amount", "unit_price",
         "vendor", "department", "class_ru"
       }

Стиль:
    - Нумерованные секции и краткие комментарии, поясняющие назначение и логику работы каждой части.
"""

# 1. Импорт библиотек и настройка логирования
from pathlib import Path
from typing import List
from PySide6 import QtWidgets
from openpyxl import load_workbook
import csv
import logging

# Создаём логгер для модуля. Основная конфигурация задаётся в utils.init_logging().
logger = logging.getLogger(__name__)

# 2. Диалог сопоставления столбцов
class MappingDialog(QtWidgets.QDialog):
    """
    Поля маппинга:
        name, qty, coeff, amount, vendor, department, class_ru, stock_qty
    Если колонка не выбрана — значение будет пропущено или взято по умолчанию.
    """
    # 2.1 Инициализация
    def __init__(self, headers: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Сопоставление столбцов (Импорт смет)")
        layout = QtWidgets.QFormLayout(self)

        # 2.1.1 Списки выбора
        self.box_name = QtWidgets.QComboBox()
        self.box_qty = QtWidgets.QComboBox()
        self.box_coeff = QtWidgets.QComboBox()
        self.box_amount = QtWidgets.QComboBox()
        self.box_vendor = QtWidgets.QComboBox()
        self.box_department = QtWidgets.QComboBox()
        self.box_class = QtWidgets.QComboBox()
        # Добавляем поле для количества на складе. Позволяет указать колонку,
        # содержащую количество оборудования, имеющееся на складе у подрядчика.
        self.box_stock = QtWidgets.QComboBox()

        for b in (self.box_name, self.box_qty, self.box_coeff, self.box_amount, self.box_vendor, self.box_department, self.box_class, self.box_stock):
            b.addItems(["<не использовать>"] + headers)

        # 2.1.2 Значение по умолчанию для класса (если колонка не выбрана)
        self.box_default_class = QtWidgets.QComboBox()
        self.box_default_class.addItems(["Оборудование", "Персонал", "Логистика", "Расходник"])

        # 2.1.3 Компоновка
        layout.addRow("Наименование:", self.box_name)
        layout.addRow("Количество:", self.box_qty)
        layout.addRow("Коэффициент:", self.box_coeff)
        layout.addRow("Сумма:", self.box_amount)
        layout.addRow("Подрядчик:", self.box_vendor)
        layout.addRow("Отдел:", self.box_department)
        layout.addRow("Класс (колонка):", self.box_class)
        # Количество на складе: дополнительная колонка для информации о складских остатках
        layout.addRow("Кол-во на складе:", self.box_stock)
        layout.addRow("Класс по умолчанию:", self.box_default_class)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    # 2.2 Получение результата
    def get_mapping(self) -> dict:
        return {
            "name": self.box_name.currentText(),
            "qty": self.box_qty.currentText(),
            "coeff": self.box_coeff.currentText(),
            "amount": self.box_amount.currentText(),
            "vendor": self.box_vendor.currentText(),
            "department": self.box_department.currentText(),
            "class_ru": self.box_class.currentText(),
            "stock_qty": self.box_stock.currentText(),
            "default_class_ru": self.box_default_class.currentText(),
        }


# 3. Вспомогательная нормализация чисел
def _to_float(x) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


# 4. Импорт XLSX с маппингом
def _import_xlsx(path: Path, parent=None) -> list[dict]:
    """
    Импортирует данные из Excel.

    4.1. Открытие книги и выбор листа: Если в книге несколько листов,
    предлагается диалог выбора. Если пользователь отменяет выбор,
    возвращается пустой список.

    4.2. Поиск строки заголовков: Ищется первая строка с достаточным
    количеством непустых ячеек. Заголовки используются для маппинга.
    """
    # 4.1 Открытие книги
    wb = load_workbook(filename=path, data_only=True)
    sheet_name = wb.sheetnames[0]
    # Если листов несколько, запрашиваем выбор у пользователя
    if len(wb.sheetnames) > 1:
        try:
            item, ok = QtWidgets.QInputDialog.getItem(
                parent or QtWidgets.QApplication.activeWindow(),
                "Выбор листа",
                "Выберите лист Excel:",
                wb.sheetnames,
                0,
                False,
            )
            if not ok:
                logger.info("Пользователь отменил выбор листа при импорте сметы.")
                return []
            sheet_name = item
        except Exception as ex:
            # Логируем ошибку выбора листа и возвращаем пустой список
            logger.error("Ошибка выбора листа: %s", ex, exc_info=True)
            return []
    logger.info("Для импорта выбран лист: %s", sheet_name)
    ws = wb[sheet_name]

    headers: list[str] = []
    header_row_idx = None
    # 4.2 Поиск строки заголовков: ищем первую строку с большим числом непустых ячеек
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_empty) >= max(2, len(row) // 3):
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break
    if not headers:
        logger.error("Не удалось найти строку заголовков в Excel.")
        raise RuntimeError("Не удалось найти строку заголовков в Excel.")

    dlg = MappingDialog(headers, parent=parent)
    if dlg.exec() != QtWidgets.QDialog.Accepted:
        return []
    mapping = dlg.get_mapping()

    def idx_of(name: str):
        if name == "<не использовать>":
            return None
        try:
            return headers.index(name)
        except ValueError:
            return None

    i_name = idx_of(mapping["name"])
    i_qty = idx_of(mapping["qty"])
    i_coeff = idx_of(mapping["coeff"])
    i_amount = idx_of(mapping["amount"])
    i_vendor = idx_of(mapping["vendor"])
    i_department = idx_of(mapping["department"])
    i_class = idx_of(mapping["class_ru"])
    i_stock = idx_of(mapping.get("stock_qty", "<не использовать>"))
    default_class_ru = mapping["default_class_ru"]

    items: list[dict] = []
    for r in ws.iter_rows(min_row=(header_row_idx or 1) + 1, values_only=True):
        vals = ["" if c is None else c for c in r]
        name = (str(vals[i_name]).strip() if i_name is not None and i_name < len(vals) else "")
        if not name:
            continue

        qty = _to_float(vals[i_qty]) if i_qty is not None and i_qty < len(vals) else 1.0
        coeff = _to_float(vals[i_coeff]) if i_coeff is not None and i_coeff < len(vals) else 1.0
        amount = _to_float(vals[i_amount]) if i_amount is not None and i_amount < len(vals) else 0.0
        unit_price = amount / (qty * coeff) if qty > 0 and coeff > 0 else 0.0

        vendor = str(vals[i_vendor]).strip() if i_vendor is not None and i_vendor < len(vals) else ""
        department = str(vals[i_department]).strip() if i_department is not None and i_department < len(vals) else ""
        class_ru = str(vals[i_class]).strip() if i_class is not None and i_class < len(vals) else default_class_ru
        stock_val = _to_float(vals[i_stock]) if i_stock is not None and i_stock < len(vals) else 0.0

        items.append({
            "name": name,
            "qty": qty,
            "coeff": coeff,
            "amount": amount,
            "unit_price": unit_price,
            "vendor": vendor,
            "department": department,
            "class_ru": class_ru or default_class_ru,
            "stock_qty": stock_val,
        })
    return items


# 5. Импорт CSV с маппингом (первая строка — заголовки)
def _import_csv(path: Path, parent=None) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return []

    headers = [str(c).strip() for c in rows[0]]
    dlg = MappingDialog(headers, parent=parent)
    if dlg.exec() != QtWidgets.QDialog.Accepted:
        return []
    mapping = dlg.get_mapping()

    def idx_of(name: str):
        if name == "<не использовать>":
            return None
        try:
            return headers.index(name)
        except ValueError:
            return None

    i_name = idx_of(mapping["name"])
    i_qty = idx_of(mapping["qty"])
    i_coeff = idx_of(mapping["coeff"])
    i_amount = idx_of(mapping["amount"])
    i_vendor = idx_of(mapping["vendor"])
    i_department = idx_of(mapping["department"])
    i_class = idx_of(mapping["class_ru"])
    i_stock = idx_of(mapping.get("stock_qty", "<не использовать>"))
    default_class_ru = mapping["default_class_ru"]

    items: list[dict] = []
    for r in rows[1:]:
        vals = ["" if c is None else c for c in r]
        name = (str(vals[i_name]).strip() if i_name is not None and i_name < len(vals) else "")
        if not name:
            continue

        qty = _to_float(vals[i_qty]) if i_qty is not None and i_qty < len(vals) else 1.0
        coeff = _to_float(vals[i_coeff]) if i_coeff is not None and i_coeff < len(vals) else 1.0
        amount = _to_float(vals[i_amount]) if i_amount is not None and i_amount < len(vals) else 0.0
        unit_price = amount / (qty * coeff) if qty > 0 and coeff > 0 else 0.0

        vendor = str(vals[i_vendor]).strip() if i_vendor is not None and i_vendor < len(vals) else ""
        department = str(vals[i_department]).strip() if i_department is not None and i_department < len(vals) else ""
        class_ru = str(vals[i_class]).strip() if i_class is not None and i_class < len(vals) else default_class_ru
        stock_val = _to_float(vals[i_stock]) if i_stock is not None and i_stock < len(vals) else 0.0

        items.append({
            "name": name,
            "qty": qty,
            "coeff": coeff,
            "amount": amount,
            "unit_price": unit_price,
            "vendor": vendor,
            "department": department,
            "class_ru": class_ru or default_class_ru,
            "stock_qty": stock_val,
        })
    return items


# 6. Универсальный импорт по расширению
def import_file(path: Path, parent=None) -> list[dict]:
    """
    Возвращает список элементов для добавления в проект и/или каталог.
    Поддерживаемые расширения: .xlsx, .csv
    """
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _import_xlsx(path, parent=parent)
    elif ext == ".csv":
        return _import_csv(path, parent=parent)
    else:
        raise RuntimeError("Поддерживаются только файлы XLSX и CSV.")
