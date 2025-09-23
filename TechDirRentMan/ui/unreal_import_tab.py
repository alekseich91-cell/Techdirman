"""
Вкладка ``Импорт из Unreal Engine`` для TechDirRentMan.

Назначение
-----------

Эта вкладка позволяет импортировать Excel‑таблицы, сформированные в Unreal
Engine (UE5), и сопоставлять их строки с позициями глобального каталога.

Таблица UE обычно содержит два столбца: наименование позиции и количество.
Пользователь выбирает файл, далее импортированные строки отображаются в
таблице. Для каждой строки доступна кнопка «Привязать», открывающая
диалог выбора позиции из базы данных. После выбора наименования из
каталога оригинальное имя заменяется на выбранное, количество остаётся,
а остальные параметры (цена, подрядчик, отдел, класс и мощность) берутся
из записи каталога. По завершении сопоставления пользователь может
добавить выбранные позиции в текущий проект.

Принцип работы
--------------

* ``build_unreal_tab(page, tab)`` — создаёт интерфейс вкладки, включая
  элементы выбора файла, таблицу импортированных данных и кнопки действий.
* При выборе файла таблица заполняется списком импортированных строк
  (имя + количество). Для каждой строки создаётся кнопка «Привязать».
* ``_ue_assign_row`` — обработчик нажатия кнопки привязки. Открывает
  модальное окно выбора позиции из каталога, заполняя фильтры подрядчика
  и отдела. После выбора строка таблицы обновляется.
* ``_ue_add_to_summary`` — собирает выбранные строки и сохраняет их в
  проект через ``DB.add_items_bulk``. При необходимости дополняет
  глобальный каталог посредством ``DB.catalog_add_or_ignore``.

Стиль
-----

Код разбит на небольшие функции с краткими комментариями, поясняющими
назначение и логику работы. Для отображения ошибок и информации
используется ``page._log``.
"""

from __future__ import annotations

# 1. Импорт стандартных модулей
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import logging

# 2. Импорт библиотек PySide6
from PySide6 import QtWidgets, QtCore
from openpyxl import load_workbook

# 3. Импорт внутренних модулей
from .common import to_float, normalize_case, fmt_num, CLASS_EN2RU, CLASS_RU2EN
from .widgets import FileDropLabel

logger = logging.getLogger(__name__)


class UEMappingDialog(QtWidgets.QDialog):
    """
    Диалог выбора столбцов для импорта UE.

    Позволяет выбрать, какие столбцы Excel‑файла содержат наименование и
    количество. Пользователь выбирает соответствующие колонки из списка
    заголовков. Если заголовки отсутствуют, используются имена вида
    «Столбец 1», «Столбец 2», ….

    Метод :func:`get_mapping` возвращает кортеж из двух индексов:
    ``(i_name, i_qty)``. Если пользователь отменяет выбор, оба индекса
    равны ``None``.
    """

    def __init__(self, headers: List[str], parent: Optional[QtWidgets.QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Сопоставление столбцов (UE импорт)")
        self._i_name: Optional[int] = None
        self._i_qty: Optional[int] = None
        layout = QtWidgets.QFormLayout(self)

        self.cmb_name = QtWidgets.QComboBox()
        self.cmb_qty = QtWidgets.QComboBox()

        # Заполняем комбобоксы именами столбцов
        for idx, h in enumerate(headers):
            title = h or f"Столбец {idx + 1}"
            self.cmb_name.addItem(title, idx)
            self.cmb_qty.addItem(title, idx)

        # Предугадываем выбор: ищем ключевые слова
        for i, h in enumerate(headers):
            hlow = (h or "").strip().lower()
            # Наименование
            if self.cmb_name.currentIndex() == 0 and (
                "name" in hlow or "наимен" in hlow or "позиция" in hlow
            ):
                self.cmb_name.setCurrentIndex(i)
            # Количество
            if self.cmb_qty.currentIndex() == 0 and (
                "qty" in hlow or "кол" in hlow
            ):
                self.cmb_qty.setCurrentIndex(i)

        layout.addRow("Столбец с наименованием:", self.cmb_name)
        layout.addRow("Столбец с количеством:", self.cmb_qty)

        btns = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def get_mapping(self) -> Tuple[Optional[int], Optional[int]]:
        """Возвращает (i_name, i_qty), где ``None`` означает отсутствие выбора."""
        if self.cmb_name.currentIndex() >= 0:
            self._i_name = self.cmb_name.currentData()
        if self.cmb_qty.currentIndex() >= 0:
            self._i_qty = self.cmb_qty.currentData()
        return (self._i_name, self._i_qty)


def _read_ue_xlsx(path: Path) -> List[Dict[str, Any]]:
    """Читает Excel‑файл UE и возвращает список элементов.

    Предполагается, что первая строка содержит заголовки, поэтому
    чтение начинается со второй строки. Если заголовки отсутствуют,
    чтение продолжается, пока есть две непустые ячейки: имя и количество.

    :param path: путь к файлу Excel
    :return: список словарей с ключами name и qty
    """
    items: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(filename=path, data_only=True)
        # Используем первый лист
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Пропускаем первую строку как заголовок, если содержит текст
            if i == 0:
                continue
            # Имя
            name = ""
            qty = 0.0
            if row and len(row) >= 1:
                val = row[0]
                if val is not None:
                    name = str(val).strip()
            # Пропускаем строки без имени
            if not name:
                continue
            # Количество
            if row and len(row) >= 2:
                qty = to_float(row[1], 1.0)
            else:
                qty = 1.0
            items.append({"name": name, "qty": qty})
    except Exception as ex:
        logger.error("Ошибка чтения файла UE: %s", ex, exc_info=True)
    return items


class CatalogSelectDialog(QtWidgets.QDialog):
    """
    Диалог выбора позиции из глобального каталога.

    Позволяет искать по имени, фильтровать по подрядчику и отделу и
    выбирать одну запись. После подтверждения выбранная запись доступна
    через атрибут ``selected_row``.
    """
    def __init__(self, page: Any, parent: Optional[QtWidgets.QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Выбор позиции из базы данных")
        self.resize(800, 500)
        self.page = page
        self.selected_row: Optional[Dict[str, Any]] = None

        # Основная компоновка
        v = QtWidgets.QVBoxLayout(self)
        v.setContentsMargins(8, 8, 8, 8)
        v.setSpacing(6)

        # 4.1 Фильтры: поиск, подрядчик, отдел
        filt = QtWidgets.QHBoxLayout()
        filt.setSpacing(6)
        self.ed_search = QtWidgets.QLineEdit()
        self.ed_search.setPlaceholderText("Поиск по наименованию…")
        self.ed_search.setMinimumWidth(200)
        self.cmb_vendor = QtWidgets.QComboBox()
        self.cmb_vendor.setMinimumWidth(150)
        self.cmb_department = QtWidgets.QComboBox()
        self.cmb_department.setMinimumWidth(150)
        filt.addWidget(QtWidgets.QLabel("Поиск:"))
        filt.addWidget(self.ed_search)
        filt.addWidget(QtWidgets.QLabel("Подрядчик:"))
        filt.addWidget(self.cmb_vendor)
        filt.addWidget(QtWidgets.QLabel("Отдел:"))
        filt.addWidget(self.cmb_department)
        filt.addStretch(1)
        v.addLayout(filt)

        # 4.2 Таблица каталога
        self.tbl = QtWidgets.QTableWidget(0, 6)
        self.tbl.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        headers = ["Наименование", "Класс", "Подрядчик", "Цена", "Потр. (Вт)", "Отдел"]
        self.tbl.setHorizontalHeaderLabels(headers)
        self.tbl.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        v.addWidget(self.tbl, 1)

        # 4.3 Кнопки OK/Cancel
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        v.addWidget(btns)

        # 4.4 Заполняем фильтры и таблицу
        self._fill_filters()
        self._update_table()

        # 4.5 Сигналы
        self.ed_search.textChanged.connect(self._update_table)
        self.cmb_vendor.currentIndexChanged.connect(self._update_table)
        self.cmb_department.currentIndexChanged.connect(self._update_table)
        btns.accepted.connect(self._on_accept)
        btns.rejected.connect(self.reject)
        self.tbl.itemDoubleClicked.connect(lambda _: self._on_accept())

    def _fill_filters(self) -> None:
        """Заполняет комбобоксы подрядчиков и отделов данными из БД."""
        try:
            vendors = self.page.db.catalog_distinct_values("vendor")
            departments = self.page.db.catalog_distinct_values("department")
        except Exception as ex:
            self.page._log(f"Ошибка загрузки фильтров каталога: {ex}", "error")
            vendors, departments = [], []
        # Подрядчики
        self.cmb_vendor.blockSignals(True)
        self.cmb_vendor.clear()
        self.cmb_vendor.addItem("<Любой>", None)
        for v in vendors:
            if v:
                self.cmb_vendor.addItem(normalize_case(v), v)
        self.cmb_vendor.setCurrentIndex(0)
        self.cmb_vendor.blockSignals(False)
        # Отделы
        self.cmb_department.blockSignals(True)
        self.cmb_department.clear()
        self.cmb_department.addItem("<Любой>", None)
        for d in departments:
            if d:
                self.cmb_department.addItem(normalize_case(d), d)
        self.cmb_department.setCurrentIndex(0)
        self.cmb_department.blockSignals(False)

    def _update_table(self) -> None:
        """
        Обновляет таблицу каталога согласно текущим фильтрам.

        Поиск по имени выполняется без учёта регистра — введённый текст
        нормализуется функцией ``normalize_case``. Это обеспечивает
        регистронезависимый поиск позиций при привязке UE‑импорта и
        упрощает подбор оборудования из каталога.
        """
        raw_name = self.ed_search.text().strip()
        name = normalize_case(raw_name) if raw_name else ""
        vendor = self.cmb_vendor.currentData()
        dept = self.cmb_department.currentData()
        filters: Dict[str, Any] = {}
        if name:
            filters["name"] = name
        if vendor:
            filters["vendor"] = vendor
        if dept:
            filters["department"] = dept
        rows: List[Any] = []
        try:
            rows = self.page.db.catalog_list(filters)
        except Exception as ex:
            try:
                self.page._log(f"Ошибка запроса каталога: {ex}", "error")
            except Exception:
                pass
            rows = []
        self.tbl.setRowCount(0)
        for r in rows:
            idx = self.tbl.rowCount()
            self.tbl.insertRow(idx)
            name_norm = normalize_case(r["name"] or "")
            class_ru = CLASS_EN2RU.get((r["class"] or "equipment"), "Оборудование")
            vendor_norm = normalize_case(r["vendor"] or "")
            price = float(r["unit_price"] or 0.0)
            power = float(r["power_watts"] or 0.0)
            dept_norm = normalize_case(r["department"] or "")
            values = [
                name_norm,
                class_ru,
                vendor_norm,
                fmt_num(price, 2),
                fmt_num(power, 0),
                dept_norm,
            ]
            for col, val in enumerate(values):
                item = QtWidgets.QTableWidgetItem(str(val))
                if col == 0:
                    # Сохраняем оригинальную строку в UserRole
                    item.setData(QtCore.Qt.UserRole, dict(r))
                self.tbl.setItem(idx, col, item)
        try:
            self.tbl.resizeColumnsToContents()
        except Exception:
            pass

    def _on_accept(self) -> None:
        """Сохраняет выбранную строку и закрывает диалог."""
        row_idx = self.tbl.currentRow()
        if row_idx < 0:
            QtWidgets.QMessageBox.information(self, "Внимание", "Выберите позицию.")
            return
        item = self.tbl.item(row_idx, 0)
        if item:
            data = item.data(QtCore.Qt.UserRole)
            if data:
                self.selected_row = data
        self.accept()


def build_unreal_tab(page: Any, tab: QtWidgets.QWidget) -> None:
    """Создаёт интерфейс вкладки «Импорт из UE».

    :param page: объект ProjectPage для доступа к БД и логированию
    :param tab: виджет, в который будут размещены элементы
    """
    # 5.1 Инициализируем состояние
    page._ue_items: List[Dict[str, Any]] = []  # Список импортированных строк

    root = QtWidgets.QVBoxLayout(tab)

    # 5.2 Верхняя панель выбора файла
    top = QtWidgets.QHBoxLayout()
    # Используем FileDropLabel для перетаскивания файлов
    page.ue_drop_label = FileDropLabel(accept_exts=(".xlsx", ".xls"), on_file=lambda p: _ue_on_file_selected(page, p))
    page.ue_drop_label.setMinimumHeight(48)
    page.ue_choose_btn = QtWidgets.QPushButton("Выбрать файл (UE XLSX)")
    page.ue_choose_btn.clicked.connect(lambda: _ue_choose_file(page))
    top.addWidget(page.ue_drop_label, 1)
    top.addWidget(page.ue_choose_btn)
    root.addLayout(top)

    # 5.3 Таблица для отображения импортированных и сопоставленных данных
    page.tbl_ue = QtWidgets.QTableWidget(0, 9)
    page.tbl_ue.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
    page.tbl_ue.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
    page.tbl_ue.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
    headers = [
        "Импортируемое название",  # 0
        "Кол-во",                 # 1
        "Привязка",              # 2 (кнопка)
        "Название (БД)",         # 3
        "Цена/шт",               # 4
        "Подрядчик",             # 5
        "Отдел",                 # 6
        "Класс",                 # 7
        "Потр. (Вт)"             # 8
    ]
    page.tbl_ue.setHorizontalHeaderLabels(headers)
    page.tbl_ue.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
    page.tbl_ue.horizontalHeader().setStretchLastSection(True)
    root.addWidget(page.tbl_ue, 1)

    # 5.4 Нижняя панель действий
    bottom = QtWidgets.QHBoxLayout()
    page.ue_btn_add = QtWidgets.QPushButton("Добавить в смету")
    page.ue_btn_clear = QtWidgets.QPushButton("Очистить")
    bottom.addStretch(1)
    bottom.addWidget(page.ue_btn_clear)
    bottom.addWidget(page.ue_btn_add)
    root.addLayout(bottom)

    # 5.5 Подключаем обработчики
    page.ue_btn_add.clicked.connect(lambda: _ue_add_to_summary(page))
    page.ue_btn_clear.clicked.connect(lambda: _ue_clear_table(page))

    # 5.6 Логируем создание вкладки
    if hasattr(page, "_log"):
        page._log("Вкладка «Импорт из UE» создана.")


def _ue_choose_file(page: Any) -> None:
    """Открывает диалог выбора файла UE и обрабатывает выбранный путь."""
    dlg = QtWidgets.QFileDialog(page)
    dlg.setNameFilters(["Excel файлы (*.xlsx *.xls)"])
    dlg.setFileMode(QtWidgets.QFileDialog.ExistingFile)
    if dlg.exec() != QtWidgets.QDialog.Accepted:
        return
    files = dlg.selectedFiles()
    if not files:
        return
    path = Path(files[0])
    _ue_on_file_selected(page, path)


def _ue_on_file_selected(page: Any, path: Path) -> None:
    """Обрабатывает выбор файла UE: запрашивает сопоставление столбцов и заполняет таблицу.

    При выборе файла Excel пользователь выбирает, какие столбцы содержат
    наименование и количество. Затем файл читается, и данные заполняют
    таблицу импорта.
    """
    try:
        wb = load_workbook(filename=path, data_only=True)
    except Exception as ex:
        if hasattr(page, "_log"):
            page._log(f"UE: ошибка чтения файла {path.name}: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось открыть файл: {ex}")
        return
    # Берём первый лист
    ws = wb[wb.sheetnames[0]]
    # Получаем первую строку как заголовки
    try:
        first_row = next(ws.iter_rows(values_only=True))
    except StopIteration:
        first_row = []
    headers: List[str] = []
    for idx, cell in enumerate(first_row):
        if cell is None:
            headers.append("")
        else:
            headers.append(str(cell).strip())
    # Запрашиваем сопоставление
    dlg = UEMappingDialog(headers, parent=page)
    if dlg.exec() != QtWidgets.QDialog.Accepted:
        return
    i_name, i_qty = dlg.get_mapping()
    if i_name is None or i_qty is None:
        QtWidgets.QMessageBox.information(page, "Внимание", "Не выбраны столбцы для наименования и количества.")
        return
    # Считываем строки начиная со второй (min_row=2)
    items: List[Dict[str, Any]] = []
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Получаем наименование и количество согласно выбранным столбцам
            name = ""
            qty = 0.0
            if i_name < len(row) and row[i_name] is not None:
                name = str(row[i_name]).strip()
            if not name:
                continue
            if i_qty < len(row) and row[i_qty] is not None:
                qty = to_float(row[i_qty])
            else:
                qty = 1.0
            items.append({"name": name, "qty": qty})
    except Exception as ex:
        if hasattr(page, "_log"):
            page._log(f"UE: ошибка чтения строк: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось прочитать данные: {ex}")
        return
    # Записываем импортированные элементы
    page._ue_items = []
    for it in items:
        page._ue_items.append({
            "import_name": it.get("name", ""),
            "qty": float(it.get("qty", 0.0) or 0.0),
            "catalog": None,
        })
    _ue_fill_table(page)
    if hasattr(page, "_log"):
        page._log(f"UE: загружено строк: {len(page._ue_items)} из файла {path.name}.")


def _ue_fill_table(page: Any) -> None:
    """Заполняет таблицу UE на основе импортированных элементов."""
    tbl = page.tbl_ue
    tbl.setRowCount(0)
    for idx, row in enumerate(page._ue_items):
        tbl.insertRow(idx)
        # Импортированное название
        tbl.setItem(idx, 0, QtWidgets.QTableWidgetItem(str(row["import_name"])))
        # Количество
        tbl.setItem(idx, 1, QtWidgets.QTableWidgetItem(fmt_num(row["qty"], 3)))
        # Кнопка привязки
        btn = QtWidgets.QPushButton("Привязать")
        btn.clicked.connect(lambda _, ridx=idx: _ue_assign_row(page, ridx))
        tbl.setCellWidget(idx, 2, btn)
        # Остальные ячейки пустые
        for col in range(3, tbl.columnCount()):
            tbl.setItem(idx, col, QtWidgets.QTableWidgetItem(""))
    try:
        tbl.resizeColumnsToContents()
    except Exception:
        pass


def _ue_assign_row(page: Any, row_idx: int) -> None:
    """Открывает диалог выбора позиции каталога и привязывает её к строке."""
    dlg = CatalogSelectDialog(page, parent=page)
    if dlg.exec() != QtWidgets.QDialog.Accepted:
        return
    data = dlg.selected_row
    if not data:
        return
    # Сохраняем выбранную запись
    if row_idx < 0 or row_idx >= len(page._ue_items):
        return
    page._ue_items[row_idx]["catalog"] = data
    # Обновляем отображение
    tbl = page.tbl_ue
    # Название
    name_norm = normalize_case(data.get("name", ""))
    price = float(data.get("unit_price") or 0.0)
    vendor = normalize_case(data.get("vendor") or "")
    dept = normalize_case(data.get("department") or "")
    cls = data.get("class", "equipment")
    cls_ru = CLASS_EN2RU.get(cls, "Оборудование")
    pw = float(data.get("power_watts") or 0.0)
    tbl.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(name_norm))
    tbl.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(fmt_num(price, 2)))
    tbl.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(vendor))
    tbl.setItem(row_idx, 6, QtWidgets.QTableWidgetItem(dept))
    tbl.setItem(row_idx, 7, QtWidgets.QTableWidgetItem(cls_ru))
    tbl.setItem(row_idx, 8, QtWidgets.QTableWidgetItem(fmt_num(pw, 0)))
    if hasattr(page, "_log"):
        page._log(f"UE: строка {row_idx + 1} привязана к каталогу: «{name_norm}» подрядчик «{vendor}».")


def _ue_add_to_summary(page: Any) -> None:
    """Добавляет выбранные строки UE в проект."""
    if page.project_id is None:
        QtWidgets.QMessageBox.information(page, "Внимание", "Сначала откройте проект.")
        return
    items_for_db: List[Dict[str, Any]] = []
    rows_catalog: List[Dict[str, Any]] = []
    # Формируем уникальный batch
    batch_id = f"ue-{datetime.utcnow().isoformat()}"
    for row in page._ue_items:
        catalog = row.get("catalog")
        if not catalog:
            continue  # пропускаем непривязанные строки
        qty = float(row.get("qty") or 0)
        if qty <= 0:
            continue
        # Данные каталога
        name = normalize_case(catalog.get("name") or "")
        price = float(catalog.get("unit_price") or 0.0)
        cls_en = catalog.get("class") or "equipment"
        vendor = normalize_case(catalog.get("vendor") or "")
        dept = normalize_case(catalog.get("department") or "")
        power = float(catalog.get("power_watts") or 0.0)
        items_for_db.append({
            "project_id": page.project_id,
            "type": cls_en,
            "group_name": "Аренда оборудования",
            "name": name,
            "qty": qty,
            "coeff": 1.0,
            "amount": price * qty,
            "unit_price": price,
            "source_file": "UE",
            "vendor": vendor,
            "department": dept,
            "zone": "",
            "power_watts": power,
            "import_batch": batch_id,
        })
        rows_catalog.append({
            "name": name,
            "unit_price": price,
            "class": cls_en,
            "vendor": vendor,
            "power_watts": power,
            "department": dept,
        })
    if not items_for_db:
        QtWidgets.QMessageBox.information(page, "Внимание", "Нет выбранных строк для добавления.")
        return
    try:
        # Запись в проект
        page.db.add_items_bulk(items_for_db)
        page._log(f"UE: добавлено позиций в проект: {len(items_for_db)} (batch={batch_id}).")
        # Пополнение каталога
        if rows_catalog and hasattr(page.db, "catalog_add_or_ignore"):
            page.db.catalog_add_or_ignore(rows_catalog)
            page._log(f"Каталог обновлён/проверен: {len(rows_catalog)} строк.")
    except Exception as ex:
        page._log(f"Ошибка добавления из UE: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось добавить: {ex}")
        return
    # Обновляем интерфейс
    page._reload_zone_tabs()
    QtWidgets.QMessageBox.information(page, "Готово", f"Добавлено позиций: {len(items_for_db)}")
    if hasattr(page, "_log"):
        page._log("Добавление позиций из UE завершено.")


def _ue_clear_table(page: Any) -> None:
    """Очищает таблицу и внутренний список импортированных элементов."""
    page._ue_items = []
    tbl = page.tbl_ue
    tbl.setRowCount(0)
    if hasattr(page, "_log"):
        page._log("Таблица UE очищена.")
