"""
Модуль ``import_tab`` инкапсулирует логику вкладки «Импорт смет».

Назначение:
    Ранее код импорта таблиц из XLSX/CSV размещался в ``ProjectPage``.
    Этот модуль выносит создание интерфейса и всю обработку импорта в
    отдельные функции, улучшая читаемость и позволяя сократить размер
    ``project_page.py``.

Как работает:
    • ``build_import_tab(page, tab)`` — создаёт интерфейс вкладки и
      инициализирует переменные, привязывает сигналы;
    • ``choose_file(page)`` — открывает диалог выбора файла для импорта;
    • ``on_file_dropped(page, p)`` — обработчик для перетаскивания файла;
    • ``read_source_file(page, path)`` — считывает данные из XLSX, XLS, CSV или PDF;
    • ``build_mapping_bar(page)`` — строит панель сопоставления столбцов;
    • ``fill_src_table(page)`` — заполняет таблицу предпросмотра исходных данных;
    • ``current_mapping(page)`` — возвращает словарь индекс->ключ для колонок;
    • ``rebuild_result(page)`` — собирает и отображает агрегированные данные;
    • ``update_import_button_state(page)`` — управляет доступностью кнопки «Импорт»;
    • ``refresh_vendor_dept_zone_lists(page)`` — обновляет списки подрядчиков, отделов и зон;
    • ``apply_import(page)`` — записывает собранные данные в проект и каталог;
    • ``undo_last_import(page)`` — отменяет последний импорт.

Стиль:
    • Код разбит на небольшие функции с понятными задачами.
    • Внутри функций есть комментарии, объясняющие ключевые шаги.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Tuple
from datetime import datetime

from PySide6 import QtWidgets, QtCore
import logging

# Создаём логгер для данного модуля. Настройка базового форматирования выполняется в utils.init_logging().
logger = logging.getLogger(__name__)

from .common import (
    CLASS_EN2RU, fmt_num, to_float, apply_auto_col_resize, setup_auto_col_resize,
    normalize_case
)
from .widgets import FileDropLabel
from .summary_tab import fill_manual_zone_combo, fill_manual_dept_combo  # reuse combo updates
try:
    # Диалог несоответствия мощности может отсутствовать
    from .dialogs import PowerMismatchDialog  # type: ignore
except Exception:  # noqa: B902
    PowerMismatchDialog = None  # type: ignore


# 1. Построение вкладки импорта
def build_import_tab(page: Any, tab: QtWidgets.QWidget) -> None:
    """Создаёт интерфейс вкладки «Импорт смет» и привязывает сигналы."""
    # 1.1 Переменные импорта
    page._import_file: Path | None = None
    page._src_headers: List[str] = []
    page._src_rows: List[List[Any]] = []
    page._mapping_widgets: List[QtWidgets.QComboBox] = []
    page._result_items: List[Dict[str, Any]] = []

    root = QtWidgets.QVBoxLayout(tab)

    # 1.2 Верхняя панель: выбор файла
    top1 = QtWidgets.QHBoxLayout()
    # 1.2 Верхняя панель: выбор файла
    # Разрешаем перетаскивание файлов следующих расширений: XLSX, XLS и CSV.
    # Добавлена поддержка формата XLS для старых таблиц Excel.
    # 1.2.1 Разрешаем перетаскивание файлов XLSX, XLS, CSV и PDF
    page.drop_label = FileDropLabel(accept_exts=(".xlsx", ".xls", ".csv", ".pdf"), on_file=lambda p: on_file_dropped(page, p))
    page.drop_label.setMinimumHeight(48)
    # Кнопка выбора файла. Поддерживаем XLS, XLSX и CSV.
    # Кнопка выбора файла. Теперь поддерживаем форматы XLSX, XLS, CSV и PDF.
    page.btn_choose_file = QtWidgets.QPushButton("Выбрать файл (XLSX/XLS/CSV/PDF)")
    page.btn_choose_file.clicked.connect(lambda: choose_file(page))
    top1.addWidget(page.drop_label, 1)
    top1.addWidget(page.btn_choose_file)

    # 1.3 Панель выбора подрядчика/отдела/зоны
    top2 = QtWidgets.QHBoxLayout()
    page.combo_vendor = QtWidgets.QComboBox()
    page.combo_vendor.setEditable(True)
    page.combo_vendor.setPlaceholderText("Подрядчик (обязательно)")

    page.combo_department = QtWidgets.QComboBox()
    page.combo_department.setEditable(True)
    page.combo_department.setPlaceholderText("Отдел")

    page.combo_zone = QtWidgets.QComboBox()
    page.combo_zone.setEditable(True)
    page.combo_zone.setPlaceholderText("Зона (опционально)")

    top2.addWidget(QtWidgets.QLabel("Подрядчик:"))
    top2.addWidget(page.combo_vendor, 1)
    top2.addWidget(QtWidgets.QLabel("Отдел:"))
    top2.addWidget(page.combo_department, 1)
    top2.addWidget(QtWidgets.QLabel("Зона:"))
    top2.addWidget(page.combo_zone, 1)

    # 1.3a Режим замены по подрядчику и зоне
    page.chk_replace = QtWidgets.QCheckBox("Импортировать с заменой (по подрядчику и зоне)")
    page.chk_replace.setToolTip("При включении: перед вставкой новые позиции заменят существующие позиции выбранного подрядчика в выбранной зоне.")
    top2.addWidget(page.chk_replace)

    # 1.4 Панель опций импорта
    top3 = QtWidgets.QHBoxLayout()
    page.chk_import_power = QtWidgets.QCheckBox("Импортировать потребление")
    page.combo_power_unit = QtWidgets.QComboBox()
    page.combo_power_unit.addItems(["Вт", "кВт", "А"])
    page.chk_filter_itogo = QtWidgets.QCheckBox("Отфильтровать «Итого»")
    page.chk_filter_empty = QtWidgets.QCheckBox("Убрать пустые строки")
    page.chk_filter_no_price_amount = QtWidgets.QCheckBox("Убрать строки без цены и суммы")

    top3.addWidget(page.chk_import_power)
    top3.addWidget(QtWidgets.QLabel("Ед. изм. нагрузки:"))
    top3.addWidget(page.combo_power_unit)
    top3.addStretch(1)
    top3.addWidget(page.chk_filter_itogo)
    top3.addWidget(page.chk_filter_empty)
    top3.addWidget(page.chk_filter_no_price_amount)

    # 1.5 Панель сопоставления столбцов
    page.map_scroll = QtWidgets.QScrollArea()
    page.map_scroll.setWidgetResizable(True)
    page.map_host = QtWidgets.QWidget()
    page.map_layout = QtWidgets.QHBoxLayout(page.map_host)
    page.map_layout.setContentsMargins(6, 4, 6, 4)
    page.map_layout.setSpacing(8)
    page.map_scroll.setWidget(page.map_host)
    page.map_scroll.setFixedHeight(86)

    # 1.6 Таблицы предпросмотра
    mid = QtWidgets.QHBoxLayout()
    left_v = QtWidgets.QVBoxLayout()
    right_v = QtWidgets.QVBoxLayout()

    page.tbl_src = QtWidgets.QTableWidget(0, 0)
    setup_auto_col_resize(page.tbl_src)
    page.tbl_src.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

    page.tbl_dst = QtWidgets.QTableWidget(0, 9)
    page.tbl_dst.setHorizontalHeaderLabels([
        "Наименование", "Кол-во", "Коэф.", "Цена/шт", "Сумма",
        "Потребл. (Вт)", "Класс (РУС)", "Подрядчик", "Отдел/Зона"
    ])
    setup_auto_col_resize(page.tbl_dst)

    left_v.addWidget(QtWidgets.QLabel("Исходная таблица"))
    left_v.addWidget(page.tbl_src, 1)
    right_v.addWidget(QtWidgets.QLabel("Результат импорта"))
    right_v.addWidget(page.tbl_dst, 1)

    sum_row = QtWidgets.QHBoxLayout()
    page.lbl_sum_amount = QtWidgets.QLabel("Сумма импорта: 0 ₽")
    page.lbl_sum_power = QtWidgets.QLabel("Потребление: 0 кВт")
    sum_row.addWidget(page.lbl_sum_amount)
    sum_row.addStretch(1)
    sum_row.addWidget(page.lbl_sum_power)
    right_v.addLayout(sum_row)

    mid.addLayout(left_v, 1)
    mid.addLayout(right_v, 1)

    # 1.7 Кнопки действий
    actions = QtWidgets.QHBoxLayout()
    page.btn_prepare = QtWidgets.QPushButton("Обновить предпросмотр")
    page.btn_import = QtWidgets.QPushButton("Импорт в проект")
    page.btn_import_db = QtWidgets.QPushButton("Импорт в БД")
    page.btn_undo = QtWidgets.QPushButton("Отменить импорт")
    # Кнопка импорта в БД остаётся выключенной, пока не будет выбран файл и подрядчик
    page.btn_import.setEnabled(False)
    page.btn_import_db.setEnabled(False)
    page.btn_undo.setEnabled(False)
    actions.addWidget(page.btn_prepare)
    actions.addStretch(1)
    actions.addWidget(page.btn_undo)
    actions.addWidget(page.btn_import_db)
    actions.addWidget(page.btn_import)

    # 1.8 Размещение
    root.addLayout(top1)
    root.addLayout(top2)
    root.addLayout(top3)
    root.addWidget(page.map_scroll)
    root.addLayout(mid)
    root.addLayout(actions)

    # 1.9 Сигналы
    page.chk_filter_itogo.toggled.connect(lambda: rebuild_result(page))
    page.chk_filter_empty.toggled.connect(lambda: rebuild_result(page))
    page.chk_filter_no_price_amount.toggled.connect(lambda: rebuild_result(page))
    page.chk_import_power.toggled.connect(lambda: rebuild_result(page))
    page.combo_power_unit.currentIndexChanged.connect(lambda _: rebuild_result(page))
    page.btn_prepare.clicked.connect(lambda: rebuild_result(page))
    page.btn_import.clicked.connect(lambda: apply_import(page))
    page.btn_import_db.clicked.connect(lambda: apply_import_to_catalog(page))
    page.chk_replace.toggled.connect(lambda _: update_import_button_state(page))
    page.combo_zone.editTextChanged.connect(lambda _: update_import_button_state(page))
    page.btn_undo.clicked.connect(lambda: undo_last_import(page))
    page.combo_vendor.editTextChanged.connect(lambda: update_import_button_state(page))
    page.combo_vendor.currentTextChanged.connect(lambda: update_import_button_state(page))


# 2. Диалог выбора файла
def choose_file(page: Any) -> None:
    """Открывает диалог выбора файла и передаёт результат обработчику."""
    # 2.1 Диалог выбора файла: расширен фильтр, чтобы разрешить XLS.
    path, _ = QtWidgets.QFileDialog.getOpenFileName(
        page,
        "Выберите файл XLSX/XLS/CSV/PDF",
        "",
        "Файлы (*.xlsx *.xls *.csv *.pdf)"
    )
    if not path:
        return
    on_file_dropped(page, Path(path))


# 3. Обработка перетаскивания файла
def on_file_dropped(page: Any, p: Path) -> None:
    """Устанавливает выбранный файл и инициирует чтение/предпросмотр."""
    page._import_file = p
    page._log(f"Выбран файл импорта: {p}")
    try:
        read_source_file(page, p)
        build_mapping_bar(page)
        fill_src_table(page)
        rebuild_result(page)
    except Exception as ex:
        page._log(f"Ошибка чтения файла: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось прочитать файл: {ex}")


# 4. Чтение исходного файла
def read_source_file(page: Any, path: Path) -> None:
    """Считывает таблицу из XLSX или CSV, выбирая строку заголовков автоматически."""
    ext = path.suffix.lower()
    headers: List[str] = []
    rows: List[List[Any]] = []

    # 4.1 Чтение XLSX файла. Используем openpyxl для современных форматов.
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook  # импортируем только при необходимости
        except Exception as ex:
            raise RuntimeError("Библиотека openpyxl не найдена. Установите её.") from ex

        # 4.1.1 Загружаем книгу Excel
        wb = load_workbook(filename=path, data_only=True)
        sheet_name = wb.sheetnames[0]
        # Если листов больше одного, запрашиваем выбор пользователя
        if len(wb.sheetnames) > 1:
            try:
                item, ok = QtWidgets.QInputDialog.getItem(
                    page,
                    "Выбор листа",
                    "Выберите лист Excel:",
                    wb.sheetnames,
                    0,
                    False,
                )
                if not ok:
                    logger.info("Пользователь отменил выбор листа при чтении файла")
                    raise RuntimeError("Отменён выбор листа")
                sheet_name = item
            except Exception as ex:
                logger.error("Ошибка выбора листа: %s", ex, exc_info=True)
                raise RuntimeError(f"Ошибка выбора листа: {ex}")
        # Записываем в лог и в интерфейс, какой лист выбран
        try:
            page._log(f"Выбран лист: {sheet_name}")
        except Exception:
            pass
        logger.info("Чтение данных из листа '%s'", sheet_name)
        ws = wb[sheet_name]
        header_row_idx = None
        # 4.1.2 Поиск строки заголовков: ищем строку с достаточным числом непустых ячеек
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = ["" if c is None else c for c in row]
            non_empty = [c for c in vals if str(c).strip() != ""]
            if len(non_empty) >= max(2, len(vals) // 3):
                headers = [str(c).strip() if c is not None else "" for c in vals]
                header_row_idx = i
                break
        if not headers:
            logger.error("Не удалось найти строку заголовков в Excel")
            raise RuntimeError("Не удалось найти строку заголовков в Excel.")
        # 4.1.3 Чтение остальных строк после заголовков
        for r in ws.iter_rows(min_row=(header_row_idx or 1) + 1, values_only=True):
            rows.append(["" if c is None else c for c in r])

    # 4.2 Чтение XLS файла. Для формата .xls используем pandas и xlrd.
    elif ext == ".xls":
        try:
            import pandas as pd  # импортируем только внутри обработки XLS
        except Exception as ex:
            raise RuntimeError(
                "Библиотеки pandas/xlrd не найдены. Установите их для чтения XLS."
            ) from ex

        try:
            # Читаем весь файл без заголовков. Параметр header=None оставляет первую строку как данные.
            df = pd.read_excel(path, header=None)
        except Exception as ex:
            raise RuntimeError(f"Ошибка чтения XLS: {ex}") from ex
        header_row_idx = None
        # Ищем строку заголовков: первая строка с большим количеством непустых ячеек.
        for idx, (_, row) in enumerate(df.iterrows()):
            # Заменяем NaN/None на пустую строку
            vals = ["" if (pd.isna(c) or c is None) else c for c in row]
            # Подсчитываем непустые элементы
            non_empty = [c for c in vals if str(c).strip() != ""]
            # Минимум два непустых или не менее трети всех
            if len(non_empty) >= max(2, len(vals) // 3):
                headers = [str(c).strip() if c is not None else "" for c in vals]
                header_row_idx = idx
                break
        # Если заголовок не найден, бросаем исключение
        if header_row_idx is None:
            raise RuntimeError("Не удалось найти строку заголовков в Excel.")
        # Считываем строки после заголовка
        for row_idx in range((header_row_idx or 0) + 1, len(df)):
            row_vals = df.iloc[row_idx].tolist()
            processed = ["" if (pd.isna(c) or c is None) else c for c in row_vals]
            rows.append(processed)

    elif ext == ".pdf":
        # 4.3 Чтение PDF файла. Используем pdfplumber для извлечения таблиц.
        try:
            import pdfplumber  # type: ignore
        except Exception as ex:
            raise RuntimeError(
                "Библиотека pdfplumber не найдена. Установите её для импорта PDF."
            ) from ex
        headers = []
        rows = []
        try:
            with pdfplumber.open(path) as pdf:
                for page_idx, page_obj in enumerate(pdf.pages, start=1):
                    try:
                        tables = page_obj.extract_tables() or []
                        for table in tables:
                            # Первая строка таблицы считаем заголовком
                            h, *body = table
                            if not headers:
                                headers = [str(c).strip() for c in h]
                                rows.extend(body)
                            else:
                                # Если заголовок уже есть, просто добавляем строки тела
                                rows.extend(body)
                    except Exception:
                        # Игнорируем проблемы на конкретной странице
                        continue
            if not headers:
                raise RuntimeError("В PDF не найдено таблиц для импорта.")
        except Exception as ex:
            raise RuntimeError(f"Ошибка чтения PDF: {ex}") from ex
    elif ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            table = list(reader)
        if not table:
            raise RuntimeError("CSV пустой.")
        headers = [str(c).strip() for c in table[0]]
        rows = [[c for c in r] for r in table[1:]]
    else:
        raise RuntimeError("Поддерживаются только XLSX, XLS и CSV.")

    page._src_headers = headers
    page._src_rows = rows
    page._log(f"Обнаружено столбцов: {len(headers)}; строк: {len(rows)}")


# 5. Построение панели сопоставления столбцов
def build_mapping_bar(page: Any) -> None:
    """Создаёт панель сопоставления столбцов исходной таблицы."""
    # Удаляем старые элементы
    for i in reversed(range(page.map_layout.count())):
        item = page.map_layout.takeAt(i)
        if item and item.widget():
            item.widget().deleteLater()

    page._mapping_widgets = []
    # Варианты для сопоставления
    choices = [
        ("— не использовать —", None),
        ("Наименование", "name"),
        ("Количество", "qty"),
        ("Цена", "price"),
        ("Сумма", "amount"),
        ("Коэффициент", "coeff"),
        ("Потребление", "power"),
        # Новый тип сопоставления: колонка с количеством на складе у подрядчика.
        ("Кол-во на складе", "stock_qty"),
    ]

    for idx, head in enumerate(page._src_headers):
        cell = QtWidgets.QWidget()
        v = QtWidgets.QVBoxLayout(cell)
        v.setContentsMargins(6, 2, 6, 2)
        v.setSpacing(4)

        lbl = QtWidgets.QLabel(head or f"Колонка {idx + 1}")
        lbl.setWordWrap(True)
        lbl.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        lbl.setToolTip(head or f"Колонка {idx + 1}")

        combo = QtWidgets.QComboBox()
        for title, key in choices:
            combo.addItem(title, key)

        hlow = (head or "").strip().lower()
        pre = None
        if "наимен" in hlow:
            pre = "name"
        elif "кол" in hlow:
            # Различаем количество (qty) и количество на складе (stock_qty) по ключевым словам
            # Если заголовок содержит слово "склад" или "налич", выбираем stock_qty
            if "склад" in hlow or "налич" in hlow:
                pre = "stock_qty"
            else:
                pre = "qty"
        elif "коэфф" in hlow or "коэф" in hlow:
            pre = "coeff"
        elif "сум" in hlow:
            pre = "amount"
        elif "цен" in hlow or "стоим" in hlow:
            pre = "price"
        elif "потр" in hlow or "ват" in hlow or "квт" in hlow or "ампер" in hlow:
            pre = "power"

        if pre:
            i = combo.findData(pre)
            if i >= 0:
                combo.setCurrentIndex(i)

        combo.currentIndexChanged.connect(lambda _: rebuild_result(page))
        v.addWidget(lbl)
        v.addWidget(combo)

        page.map_layout.addWidget(cell)
        page._mapping_widgets.append(combo)

    page._log("Панель сопоставления столбцов сформирована.")


# 6. Заполнение исходной таблицы
def fill_src_table(page: Any) -> None:
    """Заполняет таблицу предпросмотра исходных данных."""
    headers = page._src_headers
    rows = page._src_rows

    page.tbl_src.clear()
    page.tbl_src.setColumnCount(len(headers))
    page.tbl_src.setHorizontalHeaderLabels(headers)
    page.tbl_src.setRowCount(0)

    for r in rows:
        row = page.tbl_src.rowCount()
        page.tbl_src.insertRow(row)
        for c in range(len(headers)):
            val = r[c] if c < len(r) else ""
            page.tbl_src.setItem(row, c, QtWidgets.QTableWidgetItem(str(val)))

    apply_auto_col_resize(page.tbl_src)
    page._log(f"Предпросмотр исходной таблицы: {len(rows)} строк (автоширина применена).")


# 7. Текущее сопоставление колонок
def current_mapping(page: Any) -> Dict[int, str]:
    """Возвращает словарь {индекс: роль} для текущего сопоставления."""
    mp: Dict[int, str] = {}
    for i, cb in enumerate(page._mapping_widgets):
        role = cb.currentData()
        if role:
            mp[i] = role
    return mp


# 8. Сбор и отображение результата
def rebuild_result(page: Any) -> None:
    """Агрегирует данные по столбцам и обновляет таблицу результата."""
    if not page._src_headers:
        return

    mp = current_mapping(page)
    name_cols = [i for i, r in mp.items() if r == "name"]
    qty_cols = [i for i, r in mp.items() if r == "qty"]
    price_cols = [i for i, r in mp.items() if r == "price"]
    amount_cols = [i for i, r in mp.items() if r == "amount"]
    coeff_cols = [i for i, r in mp.items() if r == "coeff"]
    power_cols = [i for i, r in mp.items() if r == "power"]
    stock_cols = [i for i, r in mp.items() if r == "stock_qty"]

    if not name_cols:
        page._result_items = []
        page.tbl_dst.setRowCount(0)
        update_import_button_state(page)
        return

    f_itogo = page.chk_filter_itogo.isChecked()
    f_empty = page.chk_filter_empty.isChecked()
    f_no_pa = page.chk_filter_no_price_amount.isChecked()
    want_import_power = page.chk_import_power.isChecked()
    unit = page.combo_power_unit.currentText()

    agg: Dict[str, Dict[str, Any]] = {}

    for src_row in page._src_rows:
        name = ""
        if name_cols:
            c = name_cols[0]
            # Нормализуем наименование: удаляем пробелы и приводим каждое слово к Title-case
            raw_name = str(src_row[c]).strip() if c < len(src_row) and src_row[c] is not None else ""
            name = normalize_case(raw_name)

        if f_empty and not name:
            continue
        if f_itogo and ("итог" in (name or "").strip().lower()):
            continue

        qty = to_float(src_row[qty_cols[0]], 1.0) if qty_cols and qty_cols[0] < len(src_row) else 1.0
        coeff = to_float(src_row[coeff_cols[0]], 1.0) if coeff_cols and coeff_cols[0] < len(src_row) else 1.0
        price = to_float(src_row[price_cols[0]], 0.0) if price_cols and price_cols[0] < len(src_row) else 0.0
        amount = to_float(src_row[amount_cols[0]], 0.0) if amount_cols and amount_cols[0] < len(src_row) else 0.0

        if price <= 0 and amount > 0 and qty > 0 and coeff > 0:
            price = amount / (qty * coeff)
        if amount <= 0 and price > 0:
            amount = price * qty * coeff

        if f_no_pa and (price <= 0 and amount <= 0):
            continue

        power_w = 0.0
        if want_import_power and power_cols:
            pv = to_float(src_row[power_cols[0]], 0.0) if power_cols[0] < len(src_row) else 0.0
            if unit == "Вт":
                power_w = pv
            elif unit == "кВт":
                power_w = pv * 1000.0
            else:
                power_w = pv * 220.0

        if power_w <= 0:
            power_w = page.db.catalog_max_power_by_name(name) or 0.0

        # Количество на складе: если сопоставлена соответствующая колонка, берём значение, иначе 0
        stock_val: float = 0.0
        if stock_cols:
            idx = stock_cols[0]
            if idx < len(src_row):
                stock_val = to_float(src_row[idx], 0.0)

        # Используем нормализованный ключ для агрегации (нижний регистр для устойчивого сравнения)
        key = name.lower()
        if not key:
            continue

        rec = agg.get(key)
        if rec is None:
            class_en = page.db.catalog_get_class_by_name(name) or "equipment"
            agg[key] = {
                "name": name,
                "qty": qty,
                "coeff": coeff,
                "price": price,
                "amount": amount,
                "power_watts": power_w,
                "class_en": class_en,
                "stock_qty": stock_val,
            }
        else:
            rec["qty"] += qty
            rec["amount"] += amount
            if coeff > 0:
                rec["coeff"] = (rec["coeff"] + coeff) / 2.0
            if abs(power_w - rec["power_watts"]) > 1e-6:
                rec["power_watts"] = max(rec["power_watts"], power_w)
            # Суммируем количество на складе, если присутствует несколько строк
            rec["stock_qty"] = rec.get("stock_qty", 0.0) + stock_val

    page._result_items = []
    for rec in agg.values():
        qty = max(0.0, rec["qty"])
        coeff = rec["coeff"] if rec["coeff"] > 0 else 1.0
        amount = max(0.0, rec["amount"])
        price = amount / (qty * coeff) if qty > 0 and coeff > 0 else max(0.0, rec["price"])
        page._result_items.append({
            "name": rec["name"],
            "qty": qty,
            "coeff": coeff,
            "unit_price": price,
            "amount": amount,
            "power_watts": max(0.0, rec["power_watts"]),
            "class_en": rec["class_en"],
            "stock_qty": rec.get("stock_qty", 0.0),
        })

    page._result_items.sort(key=lambda x: x["name"].lower())

    page.tbl_dst.setRowCount(0)
    total_amount = 0.0
    total_power_w = 0.0

    for it in page._result_items:
        row = page.tbl_dst.rowCount()
        page.tbl_dst.insertRow(row)

        class_ru = CLASS_EN2RU.get(it["class_en"], "Оборудование")
        dept_zone = f"{page.combo_department.currentText().strip() or ''} / {page.combo_zone.currentText().strip() or ''}".strip(" /")

        vals = [
            it["name"],
            fmt_num(it["qty"], 3),
            fmt_num(it["coeff"], 3),
            fmt_num(it["unit_price"], 2),
            fmt_num(it["amount"], 2),
            fmt_num(it["power_watts"], 0),
            class_ru,
            page.combo_vendor.currentText().strip(),
            dept_zone,
        ]

        for c, v in enumerate(vals):
            page.tbl_dst.setItem(row, c, QtWidgets.QTableWidgetItem(str(v)))

        total_amount += it["amount"]
        total_power_w += it["power_watts"] * it["qty"]

    page.lbl_sum_amount.setText(f"Сумма импорта: {fmt_num(total_amount, 2)} ₽")
    page.lbl_sum_power.setText(f"Потребление: {fmt_num(total_power_w / 1000.0, 2)} кВт")
    apply_auto_col_resize(page.tbl_dst)
    page._log(
        f"Предпросмотр: позиций={len(page._result_items)}, сумма={fmt_num(total_amount,2)}, мощность={fmt_num(total_power_w/1000.0,2)} кВт."
    )
    update_import_button_state(page)


# 9. Доступность кнопки «Импорт»
def update_import_button_state(page: Any) -> None:
    """Включает или выключает кнопку импорта в зависимости от заполненности."""
    vendor_ok = bool(page.combo_vendor.currentText().strip())
    have_name_map = any(cb.currentData() == "name" for cb in page._mapping_widgets)
    enabled = vendor_ok and have_name_map and len(page._result_items) > 0
    page.btn_import.setEnabled(enabled)
    # Кнопка «Импорт в БД» должна быть активна при тех же условиях
    try:
        if getattr(page, 'btn_import_db', None) is not None:
            page.btn_import_db.setEnabled(enabled)
    except Exception:
        pass


# 10. Обновление списков подрядчиков/отделов/зон
def refresh_vendor_dept_zone_lists(page: Any) -> None:
    """Обновляет выпадающие списки подрядчиков, отделов и зон из базы данных."""
    if page.project_id is None:
        return
    vendors_raw = page.db.project_distinct_values(page.project_id, "vendor")
    depts_raw = page.db.project_distinct_values(page.project_id, "department")
    zones_raw = page.db.project_distinct_values(page.project_id, "zone")

    # Нормализуем и устраняем дубликаты без учёта регистра
    vendors = sorted({normalize_case(v) for v in vendors_raw if v}, key=lambda s: s.lower())
    depts = sorted({normalize_case(d) for d in depts_raw if d}, key=lambda s: s.lower())
    zones = sorted({normalize_case(z) for z in zones_raw if z}, key=lambda s: s.lower())

    def fill_combo(combo: QtWidgets.QComboBox, items: List[str]) -> None:
        cur = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.setEditable(True)
        for it in items:
            combo.addItem(it)
        combo.setEditText(cur)
        combo.blockSignals(False)

    fill_combo(page.combo_vendor, vendors)
    fill_combo(page.combo_department, depts)
    fill_combo(page.combo_zone, zones)

    # Обновляем выпадающие списки на панели ручного добавления в сводной смете
    fill_manual_zone_combo(page, zones)
    fill_manual_dept_combo(page, depts)


# 11. Применение импорта
def apply_import(page: Any) -> None:
    """Записывает собранные позиции в проект и каталог. Показывает диалог при несоответствии мощности."""
    if page.project_id is None or not page._result_items:
        return

    # Нормализуем подрядчика
    vendor = normalize_case(page.combo_vendor.currentText())
    if not vendor:
        QtWidgets.QMessageBox.information(page, "Внимание", "Укажите подрядчика.")
        return

    # Нормализуем отдел и зону (пустая строка допустима)
    department = normalize_case(page.combo_department.currentText())
    zone = normalize_case(page.combo_zone.currentText())

    batch_id = f"batch-{datetime.utcnow().isoformat()}"
    # 2.1 Режим замены: удаляем существующие позиции выбранного подрядчика в выбранной зоне
    replaced_backup = []
    if getattr(page, "chk_replace", None) and page.chk_replace.isChecked():
        try:
            # Пустая зона трактуется как '' (Без зоны)
            replaced_backup = page.db.delete_items_by_vendor_zone(page.project_id, vendor, zone)
            # Сообщаем об удалённых позициях (учитываем пустую зону)
            zone_str = zone or "(без зоны)"
            page._log(
                f"Замена: удалено позиций у подрядчика '{vendor}' в зоне '{zone_str}': {len(replaced_backup)}."
            )
        except Exception as ex:
            page._log(f"Ошибка удаления при замене: {ex}", "error")
            QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось удалить позиции при замене: {ex}")
            return
    # Храним резерв для undo
    page._last_replaced_items = replaced_backup

    src_file = str(page._import_file) if page._import_file else ""

    items_for_db: List[Dict[str, Any]] = []
    rows_catalog: List[Dict[str, Any]] = []
    mismatches: List[Tuple[str, str, float, List[float]]] = []

    for it in page._result_items:
        class_en = it["class_en"] or "equipment"
        pw = float(it["power_watts"] or 0)

        items_for_db.append({
            "project_id": page.project_id,
            "type": class_en,
            "group_name": "Аренда оборудования",
            "name": normalize_case(it["name"]),
            "qty": it["qty"],
            "coeff": it["coeff"],
            "amount": it["amount"],
            "unit_price": it["unit_price"],
            "source_file": src_file,
            "vendor": vendor,
            "department": department,
            "zone": zone,
            "power_watts": pw,
            "import_batch": batch_id,
        })

        rows_catalog.append({
            "name": normalize_case(it["name"]),
            "unit_price": it["unit_price"],
            "class": class_en,
            "vendor": vendor,
            "power_watts": pw,
            "department": department,
        })

        if hasattr(page.db, "catalog_distinct_powers_by_name_vendor"):
            # Для проверки различий мощности используем нормализованное имя
            olds = page.db.catalog_distinct_powers_by_name_vendor(normalize_case(it["name"]), vendor)
            olds_clean = [float(x or 0) for x in olds if x is not None]
            if pw > 0 and olds_clean and any(abs(pw - op) > 1e-6 for op in olds_clean):
                mismatches.append((it["name"], vendor, pw, olds_clean))

    if mismatches and PowerMismatchDialog is not None:
        dlg = PowerMismatchDialog(mismatches, parent=page)
        res = dlg.exec()
        if res == QtWidgets.QDialog.Accepted and hasattr(page.db, "catalog_update_power_by_name_vendor"):
            applied = 0
            for name, ven, new_pw, _olds in mismatches:
                applied += page.db.catalog_update_power_by_name_vendor(name, ven, new_pw)
            page._log(
                f"Каталог: обновлены мощности по {len(mismatches)} наименованиям, затронуто строк: {applied}."
            )
        else:
            page._log("Каталог: оставлены старые значения мощности (по запросу пользователя).")

    try:
        if items_for_db:
            page.db.add_items_bulk(items_for_db)
            page._log(f"Импортировано в проект {len(items_for_db)} позиций (batch={batch_id}).")
        if rows_catalog and hasattr(page.db, "catalog_add_or_ignore"):
            page.db.catalog_add_or_ignore(rows_catalog)
            page._log(f"Каталог пополнен/проверен, позиций: {len(rows_catalog)}.")

        # 3.2 Обновляем складские остатки для импортированных наименований, если информация о складе присутствует
        if hasattr(page.db, "catalog_update_stock_by_name_vendor"):
            for itm in page._result_items:
                try:
                    stock_q = float(itm.get("stock_qty", 0) or 0)
                except Exception:
                    stock_q = 0.0
                if stock_q > 0:
                    try:
                        # Обновляем по нормализованному имени и текущему подрядчику
                        page.db.catalog_update_stock_by_name_vendor(normalize_case(itm["name"]), vendor, stock_q)
                        page._log(
                            f"Каталог: обновлено количество на складе для «{normalize_case(itm['name'])}» подрядчик '{vendor}': {fmt_num(stock_q, 2)}",
                        )
                    except Exception as ex:
                        page._log(f"Ошибка обновления количества на складе: {ex}", "error")
    except Exception as ex:
        page._log(f"Ошибка импорта: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось провести импорт: {ex}")
        return

    page._last_import_batch = batch_id
    page.btn_undo.setEnabled(True)
    # Обновляем сводную смету
    page._reload_zone_tabs()
    QtWidgets.QMessageBox.information(page, "Готово", "Импорт завершён.")
    page._log("Импорт в проект завершён.")
    refresh_vendor_dept_zone_lists(page)


# 11.1 Импорт результатов только в базу данных (обновление каталога и складских остатков)
def apply_import_to_catalog(page: Any) -> None:
    """
    Импортирует элементы в глобальный каталог и обновляет складские остатки.

    При нажатии на кнопку «Импорт в БД» используется количество
    оборудования из результатов, чтобы заполнить столбец «Количество на
    складе» в каталоге. Цена позиции берётся из столбца «Цена/шт» (если
    доступен); никаких расчётов по формуле сумма/qty/coeff не производится.
    Позиции не добавляются в текущий проект.
    """
    if not page._result_items:
        QtWidgets.QMessageBox.information(page, "Внимание", "Нет данных для импорта в базу.")
        return
    vendor = normalize_case(page.combo_vendor.currentText())
    if not vendor:
        QtWidgets.QMessageBox.information(page, "Внимание", "Укажите подрядчика для импорта в базу.")
        return
    department = normalize_case(page.combo_department.currentText())
    rows_catalog: List[Dict[str, Any]] = []
    for it in page._result_items:
        name_norm = normalize_case(it.get("name", ""))
        if not name_norm:
            continue
        class_en = it.get("class_en", "equipment") or "equipment"
        # Используем цену /шт из результата; если отсутствует, берём 0
        unit_price = float(it.get("unit_price", 0.0) or 0.0)
        pw = float(it.get("power_watts", 0.0) or 0.0)
        rows_catalog.append({
            "name": name_norm,
            "unit_price": unit_price,
            "class": class_en,
            "vendor": vendor,
            "power_watts": pw,
            "department": department,
        })
    try:
        # Добавляем новые записи в каталог без дублей
        if rows_catalog and hasattr(page.db, "catalog_add_or_ignore"):
            page.db.catalog_add_or_ignore(rows_catalog)
            page._log(f"Каталог: пополнено/проверено {len(rows_catalog)} позиций для подрядчика '{vendor}'.")
        # Обновляем складские остатки для каждой позиции
        if hasattr(page.db, "catalog_update_stock_by_name_vendor"):
            for it in page._result_items:
                name_norm = normalize_case(it.get("name", ""))
                qty = 0.0
                try:
                    qty = float(it.get("qty", 0.0) or 0.0)
                except Exception:
                    qty = 0.0
                if not name_norm or qty <= 0:
                    continue
                try:
                    page.db.catalog_update_stock_by_name_vendor(name_norm, vendor, qty)
                    page._log(f"Каталог: обновлено количество на складе для '{name_norm}' подрядчик '{vendor}': {fmt_num(qty, 2)}")
                except Exception as ex:
                    page._log(f"Ошибка обновления склада для '{name_norm}': {ex}", "error")
        QtWidgets.QMessageBox.information(page, "Готово", f"Импорт в базу завершён, позиций: {len(rows_catalog)}")
    except Exception as ex:
        page._log(f"Ошибка импорта в базу: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось импортировать в базу: {ex}")


# 12. Отмена импорта
def undo_last_import(page: Any) -> None:
    """Удаляет позиции последнего импорта."""
    if page.project_id is None or not getattr(page, "_last_import_batch", None):
        return
    try:
        deleted = page.db.delete_items_by_import_batch(page.project_id, page._last_import_batch)
        page._log(f"Отменён импорт batch={page._last_import_batch}, удалено позиций: {deleted}.")
        QtWidgets.QMessageBox.information(page, "Готово", f"Удалено позиций: {deleted}")
    except Exception as ex:
        page._log(f"Ошибка отмены импорта: {ex}", "error")
        QtWidgets.QMessageBox.critical(page, "Ошибка", f"Не удалось отменить импорт: {ex}")
        return
    page._last_import_batch = None
    page.btn_undo.setEnabled(False)
    page._reload_zone_tabs()
